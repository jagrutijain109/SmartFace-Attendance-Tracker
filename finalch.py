from tkinter import *
from PIL import Image, ImageTk
import cv2
import face_recognition
import glob
import os
import numpy as np
from datetime import datetime
import openpyxl

# Initialize the webcam
cap = cv2.VideoCapture(0)

# Directory to save snapshots
save_dir = r'C:\Users\JAGRUTI\Desktop\face_recognition\images'

# Load known face images and names
images = []
names = []
for file in glob.glob(os.path.join(save_dir, '*.*')):
    image = cv2.imread(file)
    a = os.path.basename(file)
    b = os.path.splitext(a)[0]
    names.append(b)
    images.append(image)

# Encode known faces
encodings = [face_recognition.face_encodings(img)[0] for img in images if face_recognition.face_encodings(img)]

# Initialize unknown encodings list
unknown_encodings = []

# Initialize counter for unknown persons
unknown_count = 1

# Initialize tkinter window
win = Tk()
win.title("Face Recognition Attendance System")

label = Label(win)
label.grid()

# Global variable for name input
name_var = StringVar()

# Function to take snapshot of detected face
def snapshot(name):
    cv2image = cv2.cvtColor(cap.read()[1], cv2.COLOR_BGR2RGB)
    img = Image.fromarray(cv2image)
    face_locations = face_recognition.face_locations(cv2image)
    for face_location in face_locations:
        top, right, bottom, left = face_location
        im1 = img.crop((left, top, right, bottom))
        base_filename = os.path.join(save_dir, name)
        counter = 1
        while os.path.exists(f"{base_filename}_{counter}.jpg"):
            counter += 1
        unique_filename = f"{base_filename}_{counter}.jpg"
        im1.save(unique_filename)
        print(f"Saved: {unique_filename}")

def calculate_accuracy():
    if not names:
        return
    
    total_tests = 0  # Initialize total_tests here
    total_known = len(names)
    correct_recognized = 0
    
    for file in glob.glob(os.path.join(save_dir, '*.*')):
        total_tests += 1
        image = cv2.imread(file)
        unknown_image = face_recognition.load_image_file(file)
        unknown_encoding = face_recognition.face_encodings(unknown_image)
        
        if len(unknown_encoding) > 0:
            result = face_recognition.compare_faces(encodings, unknown_encoding[0])
            if True in result:
                correct_recognized += 1
    
    if total_known == 0:
        return
    
    recognition_accuracy = (correct_recognized / total_known) * 100
    print(f"Recognition Accuracy: {recognition_accuracy:.2f}%")


# Function to update data from saved images
def updatedata():
    global images, names
    images = []
    names = []
    for file in glob.glob(os.path.join(save_dir, '*.*')):
        image = cv2.imread(file)
        a = os.path.basename(file)
        b = os.path.splitext(a)[0]
        names.append(b)
        images.append(image)
    # Re-encode known faces
    global encodings
    encodings = [face_recognition.face_encodings(img)[0] for img in images if face_recognition.face_encodings(img)]
    print("Updated data.")

# Function to update encoded data from the current webcam view
def update_encoded_data():
    global encodings
    updatedata()
    encodings = [face_recognition.face_encodings(img)[0] for img in images if face_recognition.face_encodings(img)]


def show_frames(name_var_value):
    global unknown_encodings, unknown_count
    ret, frame = cap.read()
    if not ret:
        print("Error: Could not read frame from webcam.")
        return

    cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    img = Image.fromarray(cv2image)
    face_locations = face_recognition.face_locations(cv2image)
    face_encodings = face_recognition.face_encodings(cv2image, face_locations)

    for face_encoding, (top, right, bottom, left) in zip(face_encodings, face_locations):
        matches = face_recognition.compare_faces(encodings, face_encoding)
        name = "Unknown"

        if True in matches:
            first_match_index = matches.index(True)
            name = names[first_match_index]
        else:
            matches_unknown = face_recognition.compare_faces(unknown_encodings, face_encoding)
            if not True in matches_unknown:
                name_var_value = name_var.get()
                if name_var_value:
                    snapshot(name_var_value)
                    unknown_encodings.append(face_encoding)
                    name = name_var_value
                else:
                    snapshot(f"unknown_person_{unknown_count}")
                    unknown_encodings.append(face_encoding)
                    name = f"unknown_person_{unknown_count}"
                    unknown_count += 1
            else:
                first_match_index = matches_unknown.index(True)
                name = f"unknown_person_{first_match_index + 1}"

        main(name)

        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
        cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_DUPLEX, 0.5, (255, 255, 255), 1)

    img = Image.fromarray(frame)
    imgtk = ImageTk.PhotoImage(image=img)
    label.imgtk = imgtk
    label.configure(image=imgtk)
    label.after(10, show_frames, name_var_value)

def test_accuracy():
    total_tests = 0
    correct_recognitions = 0
    correct_attendance = 0
    total_attendance_tests = 0
    
    # Check if there are known faces to test against
    if not names:
        return
    
    # Test face recognition accuracy
    for file in glob.glob(os.path.join(save_dir, '*.*')):
        total_tests += 1
        image = cv2.imread(file)
        unknown_image = face_recognition.load_image_file(file)
        unknown_encoding = face_recognition.face_encodings(unknown_image)
        
        if len(unknown_encoding) > 0:
            result = face_recognition.compare_faces(encodings, unknown_encoding[0])
            if True in result:
                correct_recognitions += 1
    
    if total_tests == 0:
        print("No test files found. Cannot calculate recognition accuracy.")
    else:
        recognition_accuracy = (correct_recognitions / total_tests) * 100
        print(f"Recognition Accuracy: {recognition_accuracy:.2f}%")
    
    # Test attendance recording accuracy
    current_date = datetime.now().strftime("%Y-%m-%d")
    try:
        workbook = openpyxl.load_workbook(f"{current_date}.xlsx")
        sheet = workbook.active
        
        for row in sheet.iter_rows(values_only=True):
            total_attendance_tests += 1
            # Check if attendance record is correctly marked
            username = row[0]
            current_date = row[1]
            current_time = row[2]
            mark = row[3]
            
            # Perform validation based on your criteria
            if is_username_unique(sheet, username):
                # Check if mark (late/on time) is correctly recorded
                if mark in ["Late", "On Time"]:
                    correct_attendance += 1
            else:
                # Handle duplicate usernames or other scenarios
                pass
                
    except FileNotFoundError:
        print(f"File {current_date}.xlsx not found.")
    
    if total_attendance_tests == 0:
        print("No attendance records found. Cannot calculate attendance accuracy.")
    else:
        attendance_accuracy = (correct_attendance / total_attendance_tests) * 100
        print(f"Attendance Accuracy: {attendance_accuracy:.2f}%")

# Example usage in your main code


# Function to create or open workbook
def create_or_open_workbook():
    current_date = datetime.now().strftime("%Y-%m-%d")
    try:
        workbook = openpyxl.load_workbook(f"{current_date}.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.save(f"{current_date}.xlsx")
    sheet = workbook.active
    if sheet.max_row == 1:
        sheet.append(["Username", "Current Date", "Current Time", "Mark"])
    return workbook, sheet

# Function to check if username is unique
def is_username_unique(sheet, username):
    for row in sheet.iter_rows(values_only=True):
        if row[0] == username:
            return False
    return True

# Function to determine if user is late or on time
def is_late_or_ontime(current_time_str, target_time_str, lateness_threshold_minutes):
    try:
        current_time = datetime.strptime(current_time_str, '%H:%M')
        target_time = datetime.strptime(target_time_str, '%H:%M')
        time_difference = current_time - target_time
        lateness_minutes = time_difference.total_seconds() / 60
        if lateness_minutes >= lateness_threshold_minutes:
            return "Late"
        else:
            return "On Time"
    except ValueError:
        return "Invalid time format. Use 'HH:MM' format (e.g., '13:45')."

# Function to save attendance record
def main(username):
    workbook, sheet = create_or_open_workbook()
    if is_username_unique(sheet, username):
        current_time_str = datetime.now().strftime("%H:%M")
        target_time_str = '12:00'
        lateness_threshold_minutes = 6
        result = is_late_or_ontime(current_time_str, target_time_str, lateness_threshold_minutes)
        current_date = datetime.now().strftime("%Y-%m-%d")
        current_time = datetime.now().strftime("%H:%M:%S")
        sheet.append([username, current_date, current_time, result])
        workbook.save(f"{current_date}.xlsx")
        print("Attendance Marked!!")
    else:
        print("Attendance already marked please enter the new face")

# Function to handle window close
def quitapp():
    win.destroy()
    cap.release()
    cv2.destroyAllWindows()

# Start encoding known faces from images
update_encoded_data()

ret, frame = cap.read()
if ret:
    cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    img = Image.fromarray(cv2image)
    imgtk = ImageTk.PhotoImage(image=img)
    label.imgtk = imgtk
    label.configure(image=imgtk)

# Start showing frames
show_frames(name_var.get())
test_accuracy()
calculate_accuracy()

# Run tkinter main loop
win.mainloop()



